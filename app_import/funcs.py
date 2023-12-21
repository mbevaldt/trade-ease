from .models import Ledger, ProForma, Invoice, BillLading
from . import ledger
import pandas as pd
import plotly.express as px
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def get_df_merged_prof_invoice():
    
    proforma = ProForma.objects.all().values()
    df_proforma = pd.DataFrame.from_records(proforma)
    # ['id', 'pi_ref', 'other_ref', 'proforma_date', 'port_dest', 'num_item', 
    # 'brand', 'description', 'quantity', 'ncm', 'weight', 'unit_value']

    ledger = Ledger.objects.all().values()
    df_ledger = pd.DataFrame.from_records(ledger)
    # ['id', 'proforma_id', 'proforma_ref', 'proforma_date', 'invoice_id', 
    # 'invoice_ref', 'invoice_date', 'description', 'quantity_used']

    if df_proforma.empty or df_ledger.empty:
        return pd.DataFrame()

    df_prof_invoice = df_proforma.merge(df_ledger, 
                                    how='left', 
                                    left_on='id', 
                                    right_on='proforma_id')
    # ['id_x', 'pi_ref', 'other_ref', 'proforma_date_x', 'port_dest', 
    # 'num_item', 'brand', 'description_x', 'quantity', 'ncm', 'weight', 
    # 'unit_value', 'id_y', 'proforma_id', 'proforma_ref', 'proforma_date_y', 
    # 'invoice_id', 'invoice_ref', 'invoice_date', 'des
    
    return df_prof_invoice

def make_plotly_invoice():

    invoice = Invoice.objects.all().values()
    df_invoice = pd.DataFrame.from_records(invoice)
    if df_invoice.empty: return

    df_invoice['invoice_total'] = df_invoice['quantity'] * df_invoice['unit_value']
    
    df_invoice['year_month'] = (df_invoice['invoice_date'].astype(str)
                                .apply(lambda x: x[:4]+x[5:7])) #yyyymm
    
    invoice_month = (df_invoice.groupby(['customer', 'year_month'], as_index=False)
                     .agg(invoice_total=('invoice_total', 'sum')))
    
    fig = px.bar(x=invoice_month.year_month, 
                 y=invoice_month.invoice_total, 
                 color=invoice_month.customer,
                 color_discrete_sequence = px.colors.qualitative.Prism,
                 template="plotly_white", #"simple_white"
                 title="Valores Mensais das Invoices por Cliente", 
                 labels={'x': 'Mês', 'y': 'Total $'}) # text_auto=True
    
    fig.update_layout(title_font_size=28, 
                      xaxis_title_font_size=20,
                      yaxis_title_font_size=20,
                      title_x=0.5)
    
    fig.update_traces(marker_line_width=1.5, opacity=0.7)
    
    fig.show()
    #fig.write_html('dice_visual_d6d10.html')

def export_all_excel(self):

    invoice = Invoice.objects.all().values()
    df_invoice = pd.DataFrame.from_records(invoice)

    proforma = ProForma.objects.all().values()
    df_proforma = pd.DataFrame.from_records(proforma)

    ledgers = Ledger.objects.all().values()
    df_ledger = pd.DataFrame.from_records(ledgers)

    bill = BillLading.objects.all().values()
    df_bill = pd.DataFrame.from_records(bill)

    df_balance = ledger.show_ledger_balance(self)
    df_balance.iloc[:,-3:] = df_balance.iloc[:,-3:].apply(lambda x: x.str.replace(',',''))
    #df_balance['quantity', 'quantity_used', 'balance'].apply(lambda x: x.str.replace(',',''))
    df_balance['quantity']  = pd.to_numeric(df_balance['quantity']).map('{:.0f}'.format)
    df_balance['quantity_used']  = pd.to_numeric(df_balance['quantity_used']).map('{:.0f}'.format)
    df_balance['balance']  = pd.to_numeric(df_balance['balance']).map('{:.0f}'.format)

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = BASE_DIR + '/media/trade_ease_data.xlsx' 

    with pd.ExcelWriter(excel_path) as writer:  
        if not df_invoice.empty: df_invoice.to_excel(writer, sheet_name='invoices', index=False)
        if not df_proforma.empty: df_proforma.to_excel(writer, sheet_name='proformas', index=False)
        if not df_bill.empty: df_bill.to_excel(writer, sheet_name='bills', index=False)
        if not df_ledger.empty: df_ledger.to_excel(writer, sheet_name='consumo', index=False)
        if not df_balance.empty: df_balance.to_excel(writer, sheet_name='saldo', index=False)

    wb = openpyxl.load_workbook(filename=excel_path)
    cl_background = PatternFill(start_color='AAE9FB', 
                                end_color='AAE9FB', 
                                fill_type="solid")
    
    for ws in wb.worksheets:
        ws.freeze_panes = ws['A2']

        for row in ws.iter_rows(min_row=1, max_row=1): 
            for cell in row:
                cell.fill = cl_background
                ws.column_dimensions[get_column_letter(cell.column)].bestFit = True

    wb.save(filename=excel_path)
    
    return 'trade_ease_data.xlsx'

def show_report_excel(self):

    df_balance = ledger.show_ledger_balance(self)
    df_balance = df_balance.iloc[: , 1:]
    df_balance.iloc[:,-3:] = df_balance.iloc[:,-3:].apply(lambda x: x.str.replace(',',''))
    df_balance['quantity']  = pd.to_numeric(df_balance['quantity']).map('{:.0f}'.format)
    df_balance['quantity_used']  = pd.to_numeric(df_balance['quantity_used']).map('{:.0f}'.format)
    df_balance['balance']  = pd.to_numeric(df_balance['balance']).map('{:.0f}'.format)
    df_balance = df_balance.astype({'quantity':'int', 'quantity_used':'int', 'balance':'int'})

    df_balance_tot = (df_balance.groupby(['customer_x', 'pi_ref', 
                                          'proforma_date_x', 'size_x'], 
                                          as_index=False)
                      .agg(quantity=('quantity', 'max'), 
                            quantity_used=('quantity_used', 'sum'),
                            balance=('balance', 'min')))
    
    df_balance_tot.query('balance>0', inplace=True)
    df_balance_tot.sort_values(['proforma_date_x', 'size_x'],
                                ascending = [True, True])    
                                          
    df_balance_tyre = (df_balance_tot.groupby(['customer_x', 'size_x'], as_index=False)
                      .agg(balance=('balance', 'sum'))
                        .sort_values(['customer_x', 'size_x', 'balance'],
                                      ascending = [True, True, False]))
    df_balance_tyre.sort_values('balance', ascending=False, inplace=True)

    col_names = dict(balance='Saldo', customer_x='Cliente', invoice_date='Data', 
                     invoice_ref='Invoice', pi_ref='Pro Forma', proforma_date_x='Data', 
                     quantity='Qtdade', quantity_used='Usado', size_x='Pneu')
    
    df_balance.rename(columns=col_names, inplace=True)
    df_balance_tot.rename(columns=col_names, inplace=True)
    df_balance_tyre.rename(columns=col_names, inplace=True)
    
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = BASE_DIR + '/media/trade_ease_reports.xlsx' 

    with pd.ExcelWriter(excel_path) as writer:  
        if not df_balance.empty: df_balance.to_excel(writer, sheet_name='saldo_detalhado', index=False)
        if not df_balance_tot.empty: df_balance_tot.to_excel(writer, sheet_name='saldo_resumido', index=False)
        if not df_balance_tyre.empty: df_balance_tyre.to_excel(writer, sheet_name='saldo_pneu', index=False)

    wb = openpyxl.load_workbook(filename=excel_path)
    cl_background = PatternFill(start_color='AAE9FB', 
                                end_color='AAE9FB', 
                                fill_type="solid")
    
    for ws in wb.worksheets:
        ws.freeze_panes = ws['A2']

        for row in ws.iter_rows(min_row=1, max_row=1): 
            for cell in row:
                cell.fill = cl_background
                ws.column_dimensions[get_column_letter(cell.column)].bestFit = True

    wb.save(filename=excel_path)
    
    return 'trade_ease_reports.xlsx'