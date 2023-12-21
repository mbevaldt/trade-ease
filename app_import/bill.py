import os
import glob
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from .models import BillLading, Invoice, ProForma, Ledger
from . import ledger

def get_bills_imported():
    bills = BillLading.objects.values('bill_number').distinct()
    #QuerySet [{'bill_number': 'RSLCMBITP285'}]
    return bills

def check_bill_to_create_docs(bill_number_selected):

    bill = BillLading.objects.filter(bill_number=bill_number_selected).values()
    invoices_bill = bill[0]['invoices']
    invoices_bill = invoices_bill.split(';')
    freight = bill[0]['freight'] 
    container_bill = [reg['container'] for reg in bill]

    lst_inv, lst_cont, lst_weight, lst_msg_all = ([] for i in range(4))
    dict_msg={}

    invoice_records = Invoice.objects.all().values()
    df_invoice = pd.DataFrame.from_records(invoice_records) 
        
    if df_invoice.empty:
        dict_msg['erro_total'] = f'Nenhuma Invoice da BL{bill_number_selected} importada!'
        dict_msg['erro'] = 'Sim'

    else:
        df_invoice['invoice_base'] = df_invoice['invoice_number'].str[:10]
        df_invoice = df_invoice[df_invoice['invoice_base'].isin(invoices_bill)]
        invoices_bd = df_invoice['invoice_number'].unique()
        container_bd = df_invoice['container_num'].unique()

        df_invoice = df_invoice[df_invoice['pl_unit_weight']==0]
        lst_weight = df_invoice['invoice_number'].unique().tolist()
    
        for inv in invoices_bd:
            lst_inv.append('Sim' if inv in ','.join(invoices_bd) else 'Não')
        
        for cont in container_bill:
            lst_cont.append('Sim' if cont in ','.join(container_bd) else 'Não')
    
        df_check_inv = pd.DataFrame()
        df_check_inv['ref_number'] = invoices_bd
        df_check_inv['type'] = 'Invoice'
        df_check_inv['bd'] = lst_inv

        df_check_cont = pd.DataFrame()
        df_check_cont['ref_number'] = container_bill
        df_check_cont['type'] = 'Container'
        df_check_cont['bd'] = lst_cont  

        df_check_pl = pd.DataFrame()
        df_check_pl['ref_number'] = lst_weight
        df_check_pl['type'] = 'PL Weight'
        df_check_pl['bd'] = ['Não'] * len(lst_weight)
    
        df_check = pd.concat([df_check_inv, df_check_cont, df_check_pl])
        
        df_check = df_check[df_check['bd'] == 'Não']
        erro = 'Não' if df_check.empty else 'Sim'
        
        dict_msg['erro'] = erro
        dict_msg['bill_number'] = bill_number_selected
        dict_msg['df'] = df_check # {k: v for k, v in dict_msg.items() if v == 'Não'}
        dict_msg['freight'] = freight

    if not ProForma.objects.all():
        dict_msg['erro_total'] = 'Nenhuma Proforma importada!'
        dict_msg['erro'] = 'Sim'

    else: # se estiver tudo importado, aloca as Invoices nos saldos das Proformas
        if dict_msg['erro'] == 'Não':
            for invoice in invoices_bd:
                lst_msg = ledger.register_invoices_used('self', invoice, bill[0]['customer'])
                if lst_msg: 
                    lst_msg_all.append(lst_msg)
    if lst_msg_all: 
        dict_msg['ledger'] = lst_msg_all

    return dict_msg

def get_data_to_excel(bill_number_selected):
    
    bill = BillLading.objects.filter(bill_number=bill_number_selected).values()
    df_bill = pd.DataFrame.from_records(bill)
    invoices_bill = bill[0]['invoices'].split(';')

    df_invoice = pd.DataFrame.from_records(Invoice.objects.all().values()) 
    df_invoice['invoice_base'] = df_invoice['invoice_number'].str[:10]
    df_invoice = df_invoice[df_invoice['invoice_base'].isin(invoices_bill)] 

    df_ledger = pd.DataFrame.from_records(Ledger.objects.all().values()) 
    df_ledger['invoice_base'] = df_ledger['invoice_ref'].str[:10]
    df_ledger = df_ledger[df_ledger['invoice_base'].isin(invoices_bill)] 

    pformas = df_ledger.proforma_ref.unique()
    proforma = ProForma.objects.filter(pi_ref__in=pformas).values()
    brand = proforma[0]['brand'] 

    df_goods_excel = df_ledger.merge(df_invoice, 
                                 how='left', 
                                 left_on=['invoice_ref', 'description'],
                                 right_on=['invoice_number', 'description'])    
    
    df_goods_excel['brand'] = brand
    df_goods_excel['total_weight'] = df_goods_excel['quantity_used'] * df_goods_excel['pl_unit_weight']
    df_goods_excel['total_amount'] = df_goods_excel['quantity_used'] * df_goods_excel['unit_value']

    return df_bill, df_goods_excel

def export_to_excel(bill_number):

    CONT = 10 # number of blanks rows to list containers
    GOODS_CI = 20 # number of blanks rows to list goods in CI
    GOODS_PL = 50 # number of blanks rows to list goods in PL

    df_bill, df_goods_excel = get_data_to_excel(bill_number)

    dict_excel={}
    dict_excel['bill_number'] = bill_number
    dict_excel['min_dt_invoice'] = df_goods_excel['invoice_date_x'].min().strftime('%d.%m.%Y')
    dict_excel['proformas_used'] = ', '.join(sorted(df_goods_excel['proforma_ref'].unique()))
    dict_excel['invoices_bill'] = ', '.join(sorted(df_goods_excel['invoice_ref'].unique()))
    dict_excel['containers_list'] = sorted(df_bill['container'].values.tolist())
    dict_excel['qtd_cont'] = len(dict_excel['containers_list'])
    dict_excel['net_weight'] = f"{df_goods_excel['total_weight'].sum():,.2f}" 
    dict_excel['total_pieces'] = df_bill['quantity'].sum() 
    dict_excel['port_dest'] = df_bill.loc[0, 'port']
    dict_excel['vessel_name'] = df_bill.loc[0, 'vessel_name']
    dict_excel['freight_value'] = df_bill.loc[0, 'freight']
    dict_excel['ci_description_goods'] = 'function'
    dict_excel['pl_description_goods'] = 'function'
    
    model_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/media/'
    wb = openpyxl.load_workbook(model_path + 'CI-PL.xlsx')
    
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=150, min_col=1, max_col=15): 
            for cell in row:
                if cell.value is not None:
                    for key, val in dict_excel.items(): 
                        if key in str(cell.value): # key may be a substring of a cell value
                            
                            if key == 'containers_list':
                                _write_container_list(dict_excel, ws, cell, CONT)
                            
                            if key == 'ci_description_goods':
                                _write_goods_list_ci(df_goods_excel, ws, cell, GOODS_CI)
                            
                            if key == 'pl_description_goods':
                                _write_goods_list_pl(df_goods_excel, ws, cell, GOODS_PL)

                            else:                  
                                cell.value = str(cell.value).replace(key, str(val))  

                            continue

    file_path = model_path + bill_number + '.xlsx'
    
    files = glob.glob(model_path + '*.xlsx')
    for f in files:
        if 'CI-PL' not in f: os.remove(f)

    wb.save(file_path)
    wb.close()

    return f'{bill_number}.xlsx'

def _write_container_list(dict_excel, ws, cell, CONT):
    
    c = len(dict_excel['containers_list']) 

    for r in range(CONT+1): # blanks row in the model
        if r < c: # write container number to cell
            ws.cell(row=(cell.row + r), 
                    column=cell.column, 
                    value=dict_excel['containers_list'][r])
        else:
            if (r>=5 and r>=c): # hide blanks rows
                ws.row_dimensions[cell.row + r].hidden = True

def _write_goods_list_ci(df_goods_excel, ws, cell, GOODS_CI):

    df_goods_ci = df_goods_excel[['brand',  'proforma_ref', 'description', 
                                  'quantity_used', 'unit_value', 'total_amount']]
    
    df_goods_ci = (df_goods_ci.groupby(['brand', 'proforma_ref', 'description'], as_index=False)
                                .agg(quantity_used=('quantity_used', 'sum'), 
                                     unit_value=('unit_value', 'mean'),
                                     total_amount=('total_amount', 'sum')))
    
    df_goods_ci = df_goods_ci.sort_values(['proforma_ref', 'description'],
                                          ascending = [True, True])
    n = df_goods_ci.shape[0]
    p = df_goods_ci['proforma_ref'].unique().size
    n = n + p  # adiciona linhas que vai conter apenas o numero da proforma
    df = 0 # controla a linha do dataframe que está sendo escrita
    proforma = True # controla quando deve escrever só o numero da proforma
    first_hidden = True # controla top border da primeira celula oculta

    for r in range(GOODS_CI+1): # blanks row in the model
        if r < n: # write list of goods to cell
            
            if r==0 or (proforma and df_goods_ci.iloc[df,1] != df_goods_ci.iloc[df-1,1]):
                # se mudou proforma, escreve apenas o numero da proforma
                
                _make_top_border(ws, cell, r, 8)
                
                ws.cell(row=(cell.row + r), column=(cell.column + 1)).alignment = Alignment(horizontal='left')
                ws.cell(row=(cell.row + r), column=(cell.column + 1), value = df_goods_ci.iloc[df,1])
                ws.cell(row=(cell.row + r), column=(cell.column), value = '')
                ws.cell(row=(cell.row + r), column=(cell.column + 7), value = '')
                proforma = False

            else:
                ws.cell(row=(cell.row + r), column=(cell.column), value = df_goods_ci.iloc[df,0])
                ws.cell(row=(cell.row + r), column=(cell.column + 1), value = df_goods_ci.iloc[df,2])
                ws.cell(row=(cell.row + r), column=(cell.column + 5), value = df_goods_ci.iloc[df,3])
                ws.cell(row=(cell.row + r), column=(cell.column + 6), value = df_goods_ci.iloc[df,4])
                proforma = True
                df += 1
                
        else:# hide blanks rows
            ws.row_dimensions[cell.row + r].hidden = True

            if first_hidden:
                _make_top_border(ws, cell, r, 8)
                first_hidden=False
            
def _write_goods_list_pl(df_goods_excel, ws, cell, GOODS_PL):

    df_goods_pl = df_goods_excel[['brand', 'container_num',  'proforma_ref', 
                                  'description', 'quantity_used', 'total_weight']]
    
    df_goods_pl = (df_goods_pl.groupby(['brand', 'container_num', 'proforma_ref', 
                                        'description'], as_index=False)
                                .agg(quantity_used=('quantity_used', 'sum'), 
                                     total_weight=('total_weight', 'sum')))
    
    df_goods_pl = df_goods_pl.sort_values(['container_num', 'proforma_ref', 'description'],
                                          ascending = [True, True, True])
    n = df_goods_pl.shape[0]
    first_hidden=True

    for r in range(GOODS_PL+1): # blanks row in the model

        if r < n: # write list of goods to cell

            if df_goods_pl.iloc[r,1] != df_goods_pl.iloc[r-1,1]: # mudou container
                _make_top_border(ws, cell, r, 8)

            for c in [0, 1, 2, 3, 5, 6]: # collumn 4 is merged
                col = c if c<5 else c - 1 

                ws.cell(row=(cell.row + r), 
                        column=(cell.column + c), 
                        value = df_goods_pl.iloc[r,col])
                
        else:# hide blanks rows
            ws.row_dimensions[cell.row + r].hidden = True

            if first_hidden:
                _make_top_border(ws, cell, r, 8)
                first_hidden=False

def add_border(c, b, overwrite=False):
   
   """Add a border to a cell.
   :param c: the cell to apply the new border
   :param b: the new border of type openpyxl.styles.borders.Border
   :param overwrite: (OPTIONAL) remove existing borders on sides not defined in b (default True)
   """

   def get_border_attr(b):
       return {
          'left': getattr(b, 'left'),
          'right': getattr(b, 'right'),
          'top': getattr(b, 'top'),
          'bottom': getattr(b, 'bottom'),
      }

   if overwrite:
       c.border = b
   else:
       saved_border = get_border_attr(c.border)
       new_border = get_border_attr(b)
       c.border = Border(
           left = new_border['left'] if new_border['left'] else saved_border['left'],
           right = new_border['right'] if new_border['right'] else saved_border['right'],
           top = new_border['top'] if new_border['top'] else saved_border['top'],
           bottom = new_border['bottom'] if new_border['bottom'] else saved_border['bottom'],
      )
       
def _make_top_border(ws, cell, r, col):
    thin = Side(border_style="thin", color="000000")
    for c in range(col):
        cell_col = ws.cell(row=(cell.row + r), column=(cell.column + c))
        add_border(cell_col, Border(top=thin), overwrite=False) 










